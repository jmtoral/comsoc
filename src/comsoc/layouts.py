"""Carga del registro de layouts y detección automática del encabezado.

La detección automática es lo que hace el proyecto resistente al archivo de 2026:
la fila de encabezado se ha movido 6 -> 7 -> 8 -> 9 -> 10 -> 13 en cuatro años.
El `header_row` de layouts.yaml se usa como aserción, no como mecanismo.
"""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import openpyxl
import yaml

from .config import CONFIG_DIR, RAW_DIR
from .schema import ANCLAS_ENCABEZADO, normalizar_nombre

# Cuántas filas se escanean buscando el encabezado antes de rendirse.
MAX_FILAS_SONDEO = 30


@dataclass(frozen=True)
class Hoja:
    archivo: str
    anio: int
    generacion: str
    vintage: str
    indice: int
    nombre: str
    tipo: str  # 'polizas' | 'ejercido'
    partida_grupo: str  # '33605' | '36101-36201'
    header_row: int  # 1-indexado, esperado

    @property
    def ruta(self) -> Path:
        return RAW_DIR / self.archivo

    @property
    def id(self) -> str:
        return f"{self.anio}_{self.partida_grupo}_{self.tipo}_{self.vintage}"


@lru_cache(maxsize=1)
def cargar_layouts() -> list[Hoja]:
    with open(CONFIG_DIR / "layouts.yaml", encoding="utf-8") as fh:
        cfg = yaml.safe_load(fh)
    hojas: list[Hoja] = []
    for entrada in cfg["archivos"]:
        for h in entrada["hojas"]:
            hojas.append(
                Hoja(
                    archivo=entrada["archivo"],
                    anio=entrada["anio"],
                    generacion=entrada["generacion"],
                    vintage=entrada["vintage"],
                    indice=h["indice"],
                    nombre=h["nombre"],
                    tipo=h["tipo"],
                    partida_grupo=h["partida_grupo"],
                    header_row=h["header_row"],
                )
            )
    return hojas


def hojas_de_polizas(incluir_preliminares: bool = False) -> list[Hoja]:
    """Las hojas que forman el dataset canónico de pólizas.

    Por defecto excluye la edición preliminar de 2023 (ver PLAN_MIGRACION.md §1.3).
    """
    hojas = [h for h in cargar_layouts() if h.tipo == "polizas"]
    if not incluir_preliminares:
        hojas = [h for h in hojas if h.vintage == "definitiva"]
    return hojas


def cargar_control() -> dict[int, dict[str, dict[str, float]]]:
    """Cifras de control incrustadas en los propios archivos (2025)."""
    with open(CONFIG_DIR / "layouts.yaml", encoding="utf-8") as fh:
        cfg = yaml.safe_load(fh)
    return {e["anio"]: e["control"] for e in cfg["archivos"] if "control" in e}


def detectar_header_row(ruta: Path, nombre_hoja: str) -> int:
    """Devuelve la fila (1-indexada) del encabezado real.

    Busca la primera fila que contenga todas las ANCLAS_ENCABEZADO. Esto salta
    los bloques de título, las leyendas de partida y —en 2025— el panel de
    totales incrustado antes de la tabla.
    """
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja = libro[nombre_hoja]
        for i, fila in enumerate(hoja.iter_rows(max_row=MAX_FILAS_SONDEO, values_only=True), start=1):
            celdas = {normalizar_nombre(c) for c in fila if c is not None}
            if all(any(ancla == c for c in celdas) for ancla in ANCLAS_ENCABEZADO):
                return i
    finally:
        libro.close()
    raise ValueError(
        f"No encontré el encabezado en '{nombre_hoja}' de {ruta.name} "
        f"(busqué {ANCLAS_ENCABEZADO} en las primeras {MAX_FILAS_SONDEO} filas). "
        f"¿Cambió el formato otra vez?"
    )


def resolver_nombre_hoja(ruta: Path, hoja: Hoja) -> str:
    """Tolera cambios menores en el nombre de la pestaña.

    Primero busca el nombre exacto; si no está, cae al índice declarado y avisa.
    """
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        nombres = libro.sheetnames
    finally:
        libro.close()
    if hoja.nombre in nombres:
        return hoja.nombre
    if hoja.indice <= len(nombres):
        real = nombres[hoja.indice - 1]
        print(
            f"  [aviso] {ruta.name}: esperaba la hoja '{hoja.nombre}', "
            f"uso la #{hoja.indice} que se llama '{real}'"
        )
        return real
    raise ValueError(f"{ruta.name} no tiene hoja '{hoja.nombre}' ni índice {hoja.indice}")
